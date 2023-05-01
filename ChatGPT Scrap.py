import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from pywinauto.application import Application
import sys
from pathlib import Path
import warnings
import json
from openpyxl import Workbook

Q_and_A = list()

if getattr(sys, 'frozen', False):
    root_path = Path(sys.executable).parent
elif __file__:
    root_path = Path(__file__).parent

warnings.filterwarnings('ignore')
with open(root_path/'Config.json',encoding='utf-8') as json_file:
    Config=json.load(json_file)


username = Config["chat_gpt_user"]
password = Config["chat_gpt_pass"]
search_word = Config["search_keywords"]
BrowserHide = Config["browser_hide"]
sleep_time = Config["sleep_time"]
request_time_out_seconds = Config["time_out"]
selenium_connect_method = 2
your_liara_env_token = 'irOrJBbi64iv5hRAwwM'
liara_chrome_app_url = 'https://zargham.iran.liara.run/webdriver'
chromedriver_path = ''
# chromedriver_path = "C:/Users/zargham/.cache/selenium/chromedriver/win32/109.0.5414.74/chromedriver.exe"


if selenium_connect_method == 3:
    BrowserHide = True
# def save_to_excel():
#     path = root_path/("restaurants in "+our_query+".xlsx")
#     with pd.ExcelWriter(path, engine="openpyxl") as writer:
#         # useful code
#         df = pd.DataFrame(restaurant_list)
#         # print(df)
#         df.to_excel(writer, sheet_name='Sheet', index=False)

email = 'z.taghawi@gmail.com'
global BrowserTitleToFind
BrowserTitleToFind = 'Zargham09355124619'

URL = ""

OutPut_Filename = root_path/'results.xlsx'
workbook = Workbook()
sheet = workbook.active

sheet.cell(1,1).value='question'
sheet.cell(1,2).value='answer'
sheet.column_dimensions['A'].width=25
sheet.column_dimensions['B'].width=35

workbook.save(OutPut_Filename)

def chat_openai():
    global URL
    URL = "https://chat.openai.com/"
    my_chrome_options = webdriver.ChromeOptions()
    my_chrome_options.add_argument('--ignore-certificate-errors')
    my_chrome_options.add_argument('--ignore-ssl-errors')
    my_chrome_options.add_argument('--start-maximized')
    if BrowserHide:
        my_chrome_options.add_argument('--headless')

    # chromedriver_path = "C:/Users/zargham/.cache/selenium/chromedriver/win32/109.0.5414.74/chromedriver.exe"
    browser_driver = webdriver.Chrome(options=my_chrome_options)

    browser_driver.get(URL)
    time.sleep(20)
    sign_in_url = browser_driver.find_elements(
        By.CSS_SELECTOR, 'button.btn.relative.btn-primary:nth-child(1)')
    sign_in_url[0].click()
    time.sleep(sleep_time)
    browser_driver.execute_script('document.title = "%s"' % BrowserTitleToFind)
    time.sleep(sleep_time)
    app = Application().connect(title_re=BrowserTitleToFind)
    time.sleep(sleep_time*4)
    app.top_window().send_keystrokes(username)
    # app.top_window().send_keystrokes("{TAB 1}")
    app.top_window().send_keystrokes("{ENTER 1}")
    browser_driver.execute_script('document.title = "%s"' % BrowserTitleToFind)
    time.sleep(sleep_time)
    app = Application().connect(title_re=BrowserTitleToFind)
    time.sleep(sleep_time*4)
    app.top_window().send_keystrokes(password)
    app.top_window().send_keystrokes("{ENTER 1}")
    time.sleep(sleep_time)
    time.sleep(120)
    

# chat_openai()

def open_chatgpt():
    global URL
    my_chrome_options = webdriver.ChromeOptions()
    my_chrome_options.add_argument('--ignore-certificate-errors')
    my_chrome_options.add_argument('--ignore-ssl-errors')
    my_chrome_options.add_argument('--start-maximized')
    if BrowserHide:
        my_chrome_options.add_argument('--headless')
    URL = 'https://poe.com/ChatGPT'
    try:
        if selenium_connect_method == 1:
            if chromedriver_path != None:
                browser_driver = webdriver.Chrome(
                    chromedriver_path, options=my_chrome_options)
            else:
                browser_driver = webdriver.Chrome(options=my_chrome_options)
        elif selenium_connect_method == 2:
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.chrome.service import Service
            browser_driver = webdriver.Chrome(service=Service(
                ChromeDriverManager().install()), options=my_chrome_options)
        elif selenium_connect_method == 3:
            browser_driver = webdriver.Remote(
                command_executor=liara_chrome_app_url, options=my_chrome_options)
            # capabilities = my_chrome_options.to_capabilities()
            # browser_driver = webdriver.Remote(command_executor=liara_chrome_app_url, desired_capabilities=capabilities)
        browser_driver.get(URL)
        time.sleep(sleep_time)
    except Exception as error:
        print('---------- Exception html_format : --------\n', error)
    else:  # if try part was successful
        login_url = WebDriverWait(browser_driver, request_time_out_seconds).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, 'a.LoggedOutBotInfoPage_appButton__UO6NU')))
        login_url.click()
        time.sleep(sleep_time)
        email_sign_in = WebDriverWait(browser_driver, request_time_out_seconds).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '.Button_flat__1hj0f.undefined')))
        email_sign_in.click()
        time.sleep(sleep_time)
        browser_driver.execute_script('document.title = "%s"' % BrowserTitleToFind)
        time.sleep(sleep_time)
        app = Application().connect(title_re=BrowserTitleToFind)
        # time.sleep(sleep_time)
        app.top_window().send_keystrokes(email)
        time.sleep(sleep_time)
        app.top_window().send_keystrokes("{ENTER 1}")
        time.sleep(sleep_time)

        # send code to verify email
        # browser_driver.execute_script('document.title = "%s"' % BrowserTitleToFind)
        # time.sleep(sleep_time)
        # app = Application().connect(title_re=BrowserTitleToFind)
        # time.sleep(sleep_time)
        # app.top_window().send_keystrokes('code')
        # time.sleep(sleep_time)
        # app.top_window().send_keystrokes("{ENTER 1}")

        # log_in = WebDriverWait(browser_driver, request_time_out_seconds).until(EC.visibility_of_element_located(
        #     (By.CSS_SELECTOR, 'button.Button_buttonBase__0QP_m.Button_primary__pIDjn')))
        # log_in.click()
        # time.sleep(sleep_time)


        WebDriverWait(browser_driver, request_time_out_seconds*6).until(EC.visibility_of_all_elements_located(
            (By.CSS_SELECTOR, 'footer.ChatPageMainFooter_footer__Hm4Rt')))
        
        browser_driver.execute_script('document.title = "%s"' % BrowserTitleToFind)
        time.sleep(sleep_time)
        app = Application().connect(title_re=BrowserTitleToFind)
        # time.sleep(sleep_time)
        your_question = input('ask some question from chat-GPT: ')
        app.top_window().send_keystrokes(your_question)
        time.sleep(sleep_time)
        # app.top_window().send_keystrokes("{TAB 1}")
        app.top_window().send_keystrokes("{ENTER 1}")
        time.sleep(sleep_time)
        WebDriverWait(browser_driver, request_time_out_seconds).until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, '.ChatMessagesView_infiniteScroll__K_SeP')))
        parent_element = browser_driver.find_element(By.CSS_SELECTOR,'.ChatMessagesView_infiniteScroll__K_SeP')
        print('-- parent_elemtns.text: ',parent_element.text)
        elements = parent_element.find_elements(By.CSS_SELECTOR,'.Markdown_markdownContainer__UyYrv')
        print('-- len elements: ',len(elements))
        time.sleep(sleep_time)
        # element_number = 0
        # for element in elements:
        #     element_number += 1
        #     # print(element.text)
        #     paragraphs = element.find_elements(By.CSS_SELECTOR,'p')
        #     # print('-- len paragraph:',len(paragraphs))
        #     paragraph_number = 0
        #     for paragraph in paragraphs:
        #         paragraph_number += 1
        #         print('=== element number %i, paragraph %i, text is: %s'%(element_number,paragraph_number, paragraph.text))
        #     # if element % 2 == 0:
        #     #     Q_and_A[element/2].append([element.text])
        #     # else:
        #     #     Q_and_A.append(list(element.text))
        print('-- last answare: ',elements[-1].text)
        time.sleep(60)
open_chatgpt()
